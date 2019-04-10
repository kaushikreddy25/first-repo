'''
Created on 09-Apr-2019

@author: guruk
'''
import os
import openpyxl

os.chdir('W:\\Python\\Automate the Boring Stuff with Python Programming - Udemy\\Programs_OnlineMaterials')

workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.sheetnames)

sheet = workbook['Sheet1']

print(sheet.cell(row=1, column=2).value)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
    
print('----------------------------')
#creating on xlsx file
wb= openpyxl.Workbook()

print(wb.sheetnames)

mysheet = wb['Sheet']

mysheet.cell(row=1,column=1).value = 300
mysheet.cell(row=1,column=2).value = 50
mysheet.cell(row=1,column=3).value = 'I created this file'

os.chdir('W:\\Python\\Programs')
wb.save('CreatedFile1.xlsx')
print('File Created')

wb.create_sheet('New Sheet', index= 0)

print(wb.sheetnames)

mysheet.title= '1st sheet'
print(wb.sheetnames)

wb.save('CreatedFile2.xlsx')



